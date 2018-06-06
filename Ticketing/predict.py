import numpy as np
import xlrd,sys,io,jieba,pickle
import tensorflow as tf
from gensim.models import Word2Vec
from gensim.models.word2vec import LineSentence
from openpyxl import load_workbook
import numpy as np
import re

np.set_printoptions(suppress=True)
# fileSegWordDonePath = 'corpusSegDone.txt'
# outp1 = "outlookmodel"
# outp2 = "outlookvector.text"
# datadump = "data.pkl"
# dictdump = "dict.pkl"
sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='gb18030')

# file path need to be change
xlsfile = r"C:/Users/sesa392975/Desktop/AI study/SD Information-Mar-test.xlsx"
book = xlrd.open_workbook(xlsfile)
sheet1 = book.sheet_by_index(1)
class_nrows = sheet1.nrows

dict_index = []
lables_dict ={}
for i in range(class_nrows):
	onehot=np.zeros(class_nrows,dtype=int)
	onehot=list(onehot)
	onehot[i] = 1
	row = sheet1.row_values(i)
	class_lable = row[1]
	dict_index.append(class_lable)
	lables_dict[float(class_lable[:3])] = onehot
print(dict_index)

sheet0 = book.sheet_by_index(0)
# col0 = sheet0.col_values(0)
nrows = sheet0.nrows
wordbag = []
# lables = []
for i in range(nrows):
	row = sheet0.row_values(i)
	# lable = row[5]
	# lables.append(float(lable[:3]))
	del row[1]
	# print(i,row)
	print(i)
	wordbag.append(re.sub(r"[]\=\>\:\-\n[+\.\!\/_$-%^\\\\*(+\"\']+|[+——！。？?、~@#￥%……&*（）]+"
		, ""
		,' '.join(row)))
# print(wordbag)
wordseg = []
for i in range(len(wordbag)):
	wordseg.append([' '.join(jieba.cut(wordbag[i]))])
	if i % 100 == 0:
		print(wordseg[i])

# with open(fileSegWordDonePath,'wb') as fW:
# 	for i in range(len(wordbag)):
# 		fW.write(wordseg[i][0].encode('utf-8'))
# 		fW.write("\n".encode('utf-8'))

# print(wordseg)
# model = Word2Vec(LineSentence(fileSegWordDonePath)
# 	,size=50
# 	, window=5
# 	, min_count=5)
# model.save(outp1)
# model.wv.save_word2vec_format(outp2, binary=False)
# print(model.wv.syn0)
model = Word2Vec.load('tickets')


maxSeqLength = 50

class_dict = lables_dict
rw_data = wordseg

input_data = np.zeros((len(rw_data),maxSeqLength), dtype='int32')
count = 0
for i in rw_data:
	indexCounter = 0
	# print(i)
	for word in i[0].split(" "):
		if indexCounter < maxSeqLength:
			try:
				input_data[count,indexCounter] = model.wv.vocab[word].index
			except KeyError:
				continue
			except ValueError:
				input_data[indexCounter] = 8
		indexCounter = indexCounter + 1
	count = count + 1

numDimensions = 50 #Dimensions for each word vector
batchSize = 24
lstmUnits = 64
numClasses = class_nrows
iterations = 100000

tf.reset_default_graph()

labels = tf.placeholder(tf.float32, [batchSize, numClasses])
input_data_tf = tf.placeholder(tf.int32, [batchSize, maxSeqLength])

data = tf.Variable(tf.zeros([batchSize, maxSeqLength, numDimensions]),dtype=tf.float32)
data = tf.nn.embedding_lookup(model.wv.syn0,input_data_tf)

lstmCell = tf.contrib.rnn.BasicLSTMCell(lstmUnits)
lstmCell = tf.contrib.rnn.DropoutWrapper(cell=lstmCell, output_keep_prob=0.75)
value, _ = tf.nn.dynamic_rnn(lstmCell, data, dtype=tf.float32)

weight = tf.Variable(tf.truncated_normal([lstmUnits, numClasses]))
bias = tf.Variable(tf.constant(0.1, shape=[numClasses]))
value = tf.transpose(value, [1, 0, 2])
last = tf.gather(value, int(value.get_shape()[0]) - 1)
prediction = (tf.matmul(last, weight) + bias)

correctPred = tf.equal(tf.argmax(prediction,1), tf.argmax(labels,1))
accuracy = tf.reduce_mean(tf.cast(correctPred, tf.float32))

# loss = tf.reduce_mean(tf.nn.softmax_cross_entropy_with_logits(logits=prediction, labels=labels))
# optimizer = tf.train.AdamOptimizer().minimize(loss)

def getTestBatch(linenum):
	arr = np.zeros([batchSize, maxSeqLength])
	
	arr[0] = input_data[linenum]
	return arr

def softmax(x):
    """Compute softmax values for each sets of scores in x."""
    return np.exp(x) / np.sum(np.exp(x), axis=0)

sess = tf.InteractiveSession()
saver = tf.train.Saver()
saver.restore(sess, tf.train.latest_checkpoint('C:/Code/python_code/tickets/model'))

getwb = load_workbook(filename=xlsfile)
sheetContent = getwb[getwb.get_sheet_names()[0]]

for i in range(len(input_data)):
	#Next Batch of reviews
	nextBatch = getTestBatch(i);
	predictedSentiment = sess.run(prediction, {input_data_tf: nextBatch})[0]
	sheetContent.cell(row=i+1, column=6, value=dict_index[np.argmax(predictedSentiment)])
	sheetContent.cell(row=i+1, column=7, value=str(list(softmax(predictedSentiment))))
	# if i % 100 == 0:
	# 	print(predictedSentiment)
	# 	print(softmax(predictedSentiment))
getwb.save(filename=xlsfile)
